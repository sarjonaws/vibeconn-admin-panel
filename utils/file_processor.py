"""Procesador de archivos PDF y Excel"""
import PyPDF2
import openpyxl
from io import BytesIO

def extract_text_from_pdf(file_bytes: bytes) -> str:
    """Extrae texto de un archivo PDF"""
    pdf_file = BytesIO(file_bytes)
    pdf_reader = PyPDF2.PdfReader(pdf_file)
    
    text_parts = []
    for page in pdf_reader.pages:
        text = page.extract_text()
        if text.strip():
            text_parts.append(text)
    
    return "\n\n".join(text_parts)

def extract_text_from_excel(file_bytes: bytes) -> str:
    """Extrae texto de un archivo Excel"""
    excel_file = BytesIO(file_bytes)
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    
    text_parts = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        text_parts.append(f"=== {sheet_name} ===")
        
        for row in sheet.iter_rows(values_only=True):
            row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
            if row_text.strip():
                text_parts.append(row_text)
    
    return "\n".join(text_parts)

def process_file(filename: str, file_bytes: bytes) -> str:
    """Procesa un archivo según su extensión"""
    filename_lower = filename.lower()
    
    if filename_lower.endswith('.pdf'):
        return extract_text_from_pdf(file_bytes)
    elif filename_lower.endswith(('.xlsx', '.xls')):
        return extract_text_from_excel(file_bytes)
    else:
        raise ValueError(f"Formato no soportado: {filename}")
